import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# >>> PODEŠAVANJE <<<
FILE_PATH = "2025 Digital budget flowplan.xlsx"  # promeni ako se fajl zove drugačije
CURRENT_MONTH = 11  # 1 = Jan, 2 = Feb, ... 11 = Nov, 12 = Dec


def load_flowplan_dataframe(file_path: str) -> pd.DataFrame:
    """
    Učita tab 'V2 2025 budget digital' i pretvori ga u pandas DataFrame,
    uz normalizaciju naziva kolona za mesece (Jan, Feb, ...).
    """
    df = pd.read_excel(file_path, sheet_name="V2 2025 budget digital", header=7)

    # Ovo je red sa nazivima meseci (Jan, Feb, Mar, ...)
    month_cols_original = [
        "Q1",
        "Unnamed: 8",
        "Unnamed: 9",
        "Q2",
        "Unnamed: 11",
        "Unnamed: 12",
        "Q3",
        "Unnamed: 14",
        "Unnamed: 15",
        "Q4",
        "Unnamed: 17",
        "Unnamed: 18",
    ]

    month_names = df.iloc[0][month_cols_original].tolist()  # ['Jan', 'Feb', ... 'Dec']

    # Sklanjamo taj red (0) – on je header za mesece
    df = df.iloc[1:].reset_index(drop=True)

    # Preimenujemo kolone meseci
    rename_map = dict(zip(month_cols_original, month_names))
    df = df.rename(columns=rename_map)

    return df


def read_budgets_and_buffers(file_path: str):
    """
    Čita LTP budget i buffer za MarCom i Digital Marketing tim iz ćelija u sheet-u
    'V2 2025 budget digital':
      - MarCom LTP: K2
      - MarCom buffer: K4
      - DM LTP: Q2
      - DM buffer: Q4
    """
    wb = load_workbook(file_path, data_only=True)
    ws = wb["V2 2025 budget digital"]

    marcom_ltp = ws["K2"].value or 0
    dm_ltp = ws["Q2"].value or 0
    marcom_buffer = ws["K4"].value or 0
    dm_buffer = ws["Q4"].value or 0

    return marcom_ltp, marcom_buffer, dm_ltp, dm_buffer


def compute_team_stats(ltp: float, buffer: float, ytd_spend: float) -> dict:
    """
    Izračunava sve ključne metrike za jedan tim:
      - LTP
      - buffer
      - YTD spend
      - base_limit = LTP - buffer
      - over_vs_base = YTD - (LTP - buffer)
      - koliko LTP je ostalo / potrošeno
      - koliko buffer-a je potrošeno / ostalo
      - YTG total = max(LTP + buffer - YTD, 0)
    """
    base_limit = ltp - buffer
    over_vs_base = ytd_spend - base_limit

    # Prvo se troši LTP, tek kad je on potrošen, kreće buffer
    remaining_ltp = max(ltp - ytd_spend, 0)
    consumed_buffer = max(ytd_spend - ltp, 0)
    remaining_buffer = max(buffer - consumed_buffer, 0)

    ytg_total = max(ltp + buffer - ytd_spend, 0)

    return {
        "ltp": ltp,
        "buffer": buffer,
        "ytd_spend": ytd_spend,
        "base_limit": base_limit,
        "over_vs_base": over_vs_base,
        "remaining_ltp": remaining_ltp,
        "consumed_buffer": consumed_buffer,
        "remaining_buffer": remaining_buffer,
        "ytg_total": ytg_total,
    }


def compute_ytd_by_team(df: pd.DataFrame):
    """
    YTD trošak po timu (MarCom / Digital Marketing).

    Logika:
    - MarCom YTD = suma kolone 'Actual' (campaign header redovi)
    - DM YTD    = suma kolone 'Unnamed: 20' (DM actual) (campaign header redovi)

    "Campaign header" redovi su oni gde:
      - kolona 'CAMPAIGN' NIJE NaN
      - kolona 'Media' JE NaN
    (baš tako je strukturiran sheet koji si poslao)
    """
    # Ako iz nekog razloga kolone ne postoje, dodaj ih kao nule
    for col in ["Actual", "Unnamed: 20", "CAMPAIGN", "Media"]:
        if col not in df.columns:
            df[col] = 0

    campaign_rows = df[df["CAMPAIGN"].notna() & df["Media"].isna()]

    marcom_ytd = campaign_rows["Actual"].fillna(0).sum()
    dm_ytd = campaign_rows["Unnamed: 20"].fillna(0).sum()

    return marcom_ytd, dm_ytd


def compute_channel_spend(df: pd.DataFrame, current_month: int) -> pd.DataFrame:
    """
    Računa:
      - YTD spend po kanalu (Local Publishers, DV360, Social Media, ...)
      - Spend za aktuelni mesec po kanalu

    Ulaz:
      - df sa kolonama 'Media' i Jan..Dec
      - current_month (1-12)
    """

    month_order = [
        "Jan",
        "Feb",
        "Mar",
        "Apr",
        "May",
        "Jun",
        "Jul",
        "Aug",
        "Sep",
        "Oct",
        "Nov",
        "Dec",
    ]
    active_months = month_order[:current_month]
    current_month_name = month_order[current_month - 1]

    # Redovi gde je definisan kanal (Media kolona)
    channel_rows = df[df["Media"].notna()].copy()

    # Osiguranje da postoje kolone za sve mesece
    for m in month_order:
        if m not in channel_rows.columns:
            channel_rows[m] = 0.0

    # YTD = suma od Jan do current_month
    channel_rows["YTD"] = channel_rows[active_months].fillna(0).apply(
        pd.to_numeric, errors="coerce"
    ).fillna(0).sum(axis=1)

    # Aktuelni mesec
    channel_rows["MONTH"] = (
        pd.to_numeric(channel_rows[current_month_name], errors="coerce")
        .fillna(0)
    )

    per_channel = (
        channel_rows.groupby("Media")[["YTD", "MONTH"]].sum().reset_index()
    )

    return per_channel


def write_automated_summary(
    file_path: str,
    marcom_stats: dict,
    dm_stats: dict,
    per_channel: pd.DataFrame,
    current_month: int,
) -> str:
    """
    Kreira novi tab 'Automated Summary' sa svim traženim metrikama:
      - MarCom i Digital YTD / YTG / buffer logika
      - spend YTD po kanalu
      - spend tekući mesec po kanalu

    Fajl se snima kao <original>_automated.xlsx da ne pregaziš original.
    """
    wb = load_workbook(file_path)

    # Ako sheet već postoji, obriši ga da bismo ga ponovo kreirali
    if "Automated Summary" in wb.sheetnames:
        ws_old = wb["Automated Summary"]
        wb.remove(ws_old)

    ws = wb.create_sheet("Automated Summary")

    # ---- HEADER za timove ----
    ws["A1"] = "Team"
    ws["B1"] = "LTP"
    ws["C1"] = "Buffer"
    ws["D1"] = "YTD Spend"
    ws["E1"] = "Base limit (LTP - buffer)"
    ws["F1"] = "Over/(under) vs base"
    ws["G1"] = "Remaining LTP"
    ws["H1"] = "Consumed buffer"
    ws["I1"] = "Remaining buffer"
    ws["J1"] = "YTG total (LTP + buffer - YTD)"

    def write_team_row(row: int, name: str, stats: dict):
        ws[f"A{row}"] = name
        ws[f"B{row}"] = stats["ltp"]
        ws[f"C{row}"] = stats["buffer"]
        ws[f"D{row}"] = stats["ytd_spend"]
        ws[f"E{row}"] = stats["base_limit"]
        ws[f"F{row}"] = stats["over_vs_base"]
        ws[f"G{row}"] = stats["remaining_ltp"]
        ws[f"H{row}"] = stats["consumed_buffer"]
        ws[f"I{row}"] = stats["remaining_buffer"]
        ws[f"J{row}"] = stats["ytg_total"]

    write_team_row(2, "MarCom", marcom_stats)
    write_team_row(3, "Digital Marketing", dm_stats)

    # ---- Kanali ----
    ws["A5"] = f"Channel spend (current month = {current_month})"
    ws["A6"] = "Channel"
    ws["B6"] = "Spend YTD"
    ws["C6"] = "Spend current month"

    row = 7
    for _, rec in per_channel.sort_values("Media").iterrows():
        ws[f"A{row}"] = rec["Media"]
        ws[f"B{row}"] = float(rec["YTD"])
        ws[f"C{row}"] = float(rec["MONTH"])
        row += 1

    out_path = file_path.replace(".xlsx", "_automated.xlsx")
    wb.save(out_path)
    return out_path


def main():
    # 1) Učitavanje flowplan tabele
    df = load_flowplan_dataframe(FILE_PATH)

    # 2) LTP i buffer iz sheet-a
    marcom_ltp, marcom_buffer, dm_ltp, dm_buffer = read_budgets_and_buffers(
        FILE_PATH
    )

    # 3) YTD trošak po timu
    marcom_ytd, dm_ytd = compute_ytd_by_team(df)

    # 4) Statistike po timu (YTD, YTG, buffer logika)
    marcom_stats = compute_team_stats(marcom_ltp, marcom_buffer, marcom_ytd)
    dm_stats = compute_team_stats(dm_ltp, dm_buffer, dm_ytd)

    # 5) Trošak po kanalu (YTD + aktuelni mesec)
    per_channel = compute_channel_spend(df, CURRENT_MONTH)

    # 6) Upis u novi sheet
    out_file = write_automated_summary(
        FILE_PATH, marcom_stats, dm_stats, per_channel, CURRENT_MONTH
    )

    print("Gotovo. Napravljen fajl:", out_file)


if __name__ == "__main__":
    main()
